# analyzer_app/excel_io.py
import os
import re
import logging
from datetime import datetime, timedelta
from typing import List, Optional

import pandas as pd
from pandas.tseries.offsets import DateOffset
from openpyxl import load_workbook

from . import config
from .data_models import CalibrationStandard
from typing import Dict

logger = logging.getLogger(__name__)

def excel_coord_to_indices(coord_str: str) -> tuple[int, int]:
    """Converte una coordinata Excel (es. "B3") in indici 0-based (riga, colonna)."""
    match = re.match(r"([A-Z]+)([0-9]+)", coord_str.upper())
    if not match:
        raise ValueError(f"Coordinata Excel non valida: {coord_str}")
    col_s, row_s = match.groups()
    col_idx = 0
    for char_i, char_v in enumerate(reversed(col_s)):
        col_idx += (ord(char_v) - ord('A') + 1) * (26 ** char_i)
    return int(row_s) - 1, col_idx - 1

def parse_date_robust(date_val, context_filename: str = "N/A") -> Optional[datetime]:
    """
    Tenta di parsare una data da vari formati (stringa, timestamp, numero seriale Excel).
    """
    if pd.isna(date_val):
        return None
    if isinstance(date_val, datetime):
        return date_val
    if isinstance(date_val, pd.Timestamp):
        return date_val.to_pydatetime()

    s_date_str = str(date_val).strip()
    if not s_date_str:
        return None

    expected_formats = ['%d/%m/%Y', '%d-%m-%Y', '%d.%m.%Y', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d']
    common_formats = ['%m/%d/%Y', '%Y/%m/%d']
    all_formats_to_try = expected_formats + [f for f in common_formats if f not in expected_formats]

    for fmt in all_formats_to_try:
        try:
            return datetime.strptime(s_date_str.split(' ')[0], fmt)
        except ValueError:
            continue

    try:
        if isinstance(date_val, (int, float)) or (s_date_str.replace('.', '', 1).isdigit()):
            numeric_val = float(s_date_str)
            if 1 < numeric_val < 200000:
                return pd.to_datetime(numeric_val, unit='D', origin='1899-12-30').to_pydatetime()
    except (ValueError, TypeError, OverflowError) as e_num:
        logger.debug(f"File: {context_filename} - Parse numerico Excel fallito per '{s_date_str}': {e_num}")

    logger.warning(f"File: {context_filename} - Data '{s_date_str}' (raw: '{date_val}') non riconosciuta.")
    return None

def leggi_registro_strumenti() -> Optional[List[CalibrationStandard]]:
    # ... (Questa funzione rimane la stessa)
    if not config.FILE_REGISTRO_STRUMENTI:
        logger.error("Percorso FILE_REGISTRO_STRUMENTI non configurato. Impossibile leggere il registro.")
        return None

    logger.info(f"Tentativo lettura registro strumenti: {config.FILE_REGISTRO_STRUMENTI}")
    if not os.path.exists(config.FILE_REGISTRO_STRUMENTI):
        logger.error(f"File registro strumenti NON TROVATO: {config.FILE_REGISTRO_STRUMENTI}")
        return None

    try:
        # The order of columns in usecols must match the order in names.
        # We must sort the columns by index to ensure pandas reads them in the correct order.
        cols_to_read = {
            'modello_strumento_campione': config.REGISTRO_COL_IDX_MODELLO_STRUM_CAMPIONE,
            'id_cert_campione': config.REGISTRO_COL_IDX_ID_CERT_CAMPIONE,
            'range_campione': config.REGISTRO_COL_IDX_RANGE_CAMPIONE,
            'scadenza_cert_campione': config.REGISTRO_COL_IDX_SCADENZA_CAMPIONE
        }
        sorted_cols = sorted(cols_to_read.items(), key=lambda item: item[1])
        sorted_col_names = [item[0] for item in sorted_cols]
        sorted_col_indices = [item[1] for item in sorted_cols]

        df_registro = pd.read_excel(
            config.FILE_REGISTRO_STRUMENTI,
            sheet_name=config.REGISTRO_FOGLIO_NOME,
            header=None,
            skiprows=config.REGISTRO_RIGA_INIZIO_DATI - 1,
            usecols=sorted_col_indices,
            engine='openpyxl',
            dtype=str,
        )
        df_registro.columns = sorted_col_names

        df_registro.dropna(subset=['id_cert_campione'], inplace=True)
        df_registro = df_registro[df_registro['id_cert_campione'].astype(str).str.strip() != ""]

        strumenti_campione = []
        for _, row in df_registro.iterrows():
            id_cert_strum = str(row['id_cert_campione']).strip()
            if not id_cert_strum: continue

            scadenza_val = row['scadenza_cert_campione']
            scadenza_dt = parse_date_robust(scadenza_val, config.FILE_REGISTRO_STRUMENTI)

            data_emissione_dt = None
            if scadenza_dt:
                try: data_emissione_dt = scadenza_dt - DateOffset(years=1)
                except Exception: data_emissione_dt = scadenza_dt - timedelta(days=365)

            strumenti_campione.append(
                CalibrationStandard(
                    modello_strumento=str(row['modello_strumento_campione']).strip().upper() if not pd.isna(row['modello_strumento_campione']) else "N/D",
                    id_certificato=id_cert_strum,
                    range=str(row['range_campione']).strip() if not pd.isna(row['range_campione']) else "N/D",
                    scadenza=scadenza_dt,
                    scadenza_raw=str(scadenza_val) if not pd.isna(scadenza_val) else "N/D",
                    data_emissione=data_emissione_dt
                )
            )
        logger.info(f"Letti {len(strumenti_campione)} strumenti validi dal registro.")
        all_registry_ids = [s.id_certificato for s in strumenti_campione]
        logger.debug(f"Loaded {len(all_registry_ids)} certificate IDs from registry: {all_registry_ids}")
        return strumenti_campione
    except Exception as e:
        logger.error(f"Errore imprevisto durante lettura registro strumenti: {e}", exc_info=True)
        return None

def read_instrument_sheet_raw_data(file_path: str) -> dict:
    """
    Legge i valori grezzi da un file di scheda strumento, gestendo sia .xls che .xlsx.
    Usa openpyxl per .xlsx (veloce) e pandas+xlrd per .xls (compatibilità).
    """
    base_filename = os.path.basename(file_path)
    file_ext = os.path.splitext(base_filename)[1].lower()
    raw_data = {'file_path': file_path, 'base_filename': base_filename}

    logger.info(f"--- Inizio lettura file: {base_filename} ---")

    get_value = None
    wb = None # Per chiudere il workbook openpyxl

    if file_ext == '.xlsx':
        try:
            wb = load_workbook(filename=file_path, data_only=True, read_only=True)
            ws = wb.active
            def get_xlsx_value(coord_str):
                try:
                    cell = ws[coord_str]
                    logger.info(f"File: {base_filename} - Richiesta cella: {coord_str}")

                    for merged_range in ws.merged_cell_ranges:
                        if cell.coordinate in merged_range:
                            top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                            value = top_left_cell.value
                            logger.info(f"  -> Cella in range unito {merged_range}. Valore da '{top_left_cell.coordinate}': '{value}' (Tipo: {type(value).__name__})")
                            return value

                    value = cell.value
                    logger.info(f"  -> Cella non in range unito. Valore: '{value}' (Tipo: {type(value).__name__})")
                    return value
                except Exception as e:
                    logger.error(f"File: {base_filename} - Eccezione in get_xlsx_value per {coord_str}: {e}", exc_info=True)
                    return None
            get_value = get_xlsx_value
        except Exception as e:
            raise IOError(f"Errore apertura file .xlsx: {e}") from e

    elif file_ext == '.xls':
        try:
            df = pd.read_excel(file_path, header=None, sheet_name=0, engine='xlrd', dtype=str)
            def get_xls_value(coord_str):
                try:
                    indices = excel_coord_to_indices(coord_str)
                    return df.iloc[indices[0], indices[1]]
                except IndexError: return None
            get_value = get_xls_value
        except Exception as e:
            raise IOError(f"Errore apertura file .xls: {e}") from e
    else:
        raise ValueError(f"Formato file non supportato: {file_ext}")

    try:
        model_indicator_e2 = get_value('E2')
        model_indicator_e2_str = str(model_indicator_e2).strip().upper() if model_indicator_e2 else ""

        if "STRUMENTI DIGITALI" in model_indicator_e2_str:
            raw_data['file_type'] = "digitale"
            raw_data['sp_code'] = get_value(config.SCHEDA_DIG_CELL_TIPOLOGIA_STRUM)
            raw_data['range_um_processo'] = get_value(config.SCHEDA_DIG_CELL_RANGE_UM_PROCESSO)
            raw_data['card_date'] = get_value(config.SCHEDA_DIG_CELL_DATA_COMPILAZIONE)
            raw_data['odc'] = get_value(config.SCHEDA_DIG_CELL_ODC)
            raw_data['pdl'] = get_value(config.SCHEDA_DIG_CELL_PDL)
            raw_data['esecutore'] = get_value(config.SCHEDA_DIG_CELL_ESECUTORE)
            raw_data['supervisore'] = get_value(config.SCHEDA_DIG_CELL_SUPERVISORE_ISAB)
            raw_data['contratto'] = get_value(config.SCHEDA_DIG_CELL_CONTRATTO_COEMI)
            raw_data['cert_ids'] = [get_value(c) for c in ["C18", "E18", "G18"]]
            raw_data['cert_expiries'] = [get_value(c) for c in ["C19", "E19", "G19"]]
            raw_data['cert_models'] = [get_value(c) for c in ["C13", "E13", "G13"]]
            raw_data['cert_ranges'] = [get_value(c) for c in ["C16", "E16", "G16"]]

        elif "STRUMENTI ANALOGICI" in model_indicator_e2_str:
            raw_data['file_type'] = "analogico"
            raw_data['sp_code'] = get_value(config.SCHEDA_ANA_CELL_TIPOLOGIA_STRUM)
            raw_data['modello_l9'] = get_value(config.SCHEDA_ANA_CELL_MODELLO_STRUM)
            raw_data['card_date'] = get_value(config.SCHEDA_ANA_CELL_DATA_COMPILAZIONE)
            raw_data['range_ing'] = get_value(config.SCHEDA_ANA_CELL_RANGE_INGRESSO)
            raw_data['um_ing'] = get_value(config.SCHEDA_ANA_CELL_UM_INGRESSO)
            raw_data['range_usc'] = get_value(config.SCHEDA_ANA_CELL_RANGE_USCITA)
            raw_data['um_usc'] = get_value(config.SCHEDA_ANA_CELL_UM_USCITA)
            raw_data['range_dcs'] = get_value(config.SCHEDA_ANA_CELL_RANGE_DCS)
            raw_data['um_dcs'] = get_value(config.SCHEDA_ANA_CELL_UM_DCS)
            raw_data['odc'] = get_value(config.SCHEDA_ANA_CELL_ODC)
            raw_data['pdl'] = get_value(config.SCHEDA_ANA_CELL_PDL)
            raw_data['esecutore'] = get_value(config.SCHEDA_ANA_CELL_ESECUTORE)
            raw_data['supervisore'] = get_value(config.SCHEDA_ANA_CELL_SUPERVISORE_ISAB)
            raw_data['contratto'] = get_value(config.SCHEDA_ANA_CELL_CONTRATTO_COEMI)
            raw_data['cert_ids'] = [get_value(c) for c in ["K43", "K44", "K45"]]
            raw_data['cert_expiries'] = [get_value(c) for c in ["M43", "M44", "M45"]]
            raw_data['cert_models'] = [get_value(c) for c in ["A43", "A44", "A45"]]
            raw_data['cert_ranges'] = [get_value(c) for c in ["G43", "G44", "G45"]]
        else:
            raise ValueError(f"Tipo scheda non riconosciuto in E2: '{model_indicator_e2}' in {base_filename}")

    finally:
        if wb:
            wb.close()

    if file_ext == '.xlsx':
        logger.info(f"Dati grezzi finali estratti per {base_filename}: {raw_data}")
    logger.info(f"--- Fine lettura file: {base_filename} ---")

    return raw_data


def save_configuration(new_config: Dict[str, str]) -> bool:
    """
    Salva i nuovi percorsi di configurazione nel file parametri.xlsm.
    Restituisce True in caso di successo, False altrimenti.
    """
    try:
        wb = load_workbook(config.PATH_FILE_PARAMETRI)
        ws = wb[config.NOME_FOGLIO_PARAMETRI]

        cell_map = {
            'FILE_REGISTRO_STRUMENTI': 'B2',
            'FOLDER_PATH_DEFAULT': 'B3',
            'FILE_DATI_COMPILAZIONE_SCHEDE': 'B4',
            'FILE_MASTER_DIGITALE_XLSX': 'B5',
            'FILE_MASTER_ANALOGICO_XLSX': 'B6',
        }

        for key, cell in cell_map.items():
            if key in new_config:
                ws[cell] = new_config[key]

        wb.save(config.PATH_FILE_PARAMETRI)
        logger.info(f"Configurazione salvata con successo in {config.PATH_FILE_PARAMETRI}")
        return True
    except Exception as e:
        logger.error(f"Errore durante il salvataggio della configurazione: {e}", exc_info=True)
        return False


def write_cell(file_path: str, cell_address: str, value) -> bool:
    """
    Scrive un valore in una cella specifica di un file .xlsx.
    ATTENZIONE: Non supporta la scrittura di file .xls per preservare la formattazione.
    """
    if not file_path.lower().endswith('.xlsx'):
        logger.error(f"La scrittura è supportata solo per i file .xlsx. Impossibile modificare {os.path.basename(file_path)}")
        return False

    try:
        wb = load_workbook(file_path)
        ws = wb.active  # Assumiamo di lavorare sempre sul foglio attivo

        ws[cell_address] = value

        wb.save(file_path)
        logger.info(f"Cella {cell_address} in {os.path.basename(file_path)} aggiornata con valore '{value}'.")
        return True
    except Exception as e:
        logger.error(f"Impossibile scrivere nel file {file_path}. Errore: {e}", exc_info=True)
        return False
